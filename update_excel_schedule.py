
import datetime
from openpyxl import load_workbook
import pytz
import requests
from cogs import classes, authentication
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from dateutil.tz import *
from datetime import *

token = authentication.authenticate_WiW_API()

excel_schedule_workbook = load_workbook(authentication.get_schedule_location())
date_columns = {}
all_names = {}

all_wiw_full_names = {} # full name : status
date_row = 2

def get_url_and_headers(type):
    url = "https://api.wheniwork.com/2/" + str(type)
    headers = {
    'Host': 'api.wheniwork.com',
    'Authorization': 'Bearer ' + token, 
    }
    return [url, headers] 

def build_date_row():
    for sheet in excel_schedule_workbook.sheetnames:
        current_ws = excel_schedule_workbook[sheet]
        start_date = datetime(2022,3,2)
        for col in current_ws.iter_cols(min_row=date_row, max_row=date_row, min_col=5, max_col=500):
            for cell in col:
                cell.value = datetime.strftime(start_date, '%d %b %Y')
                date_columns[datetime.strftime(start_date, '%d %b %Y')] = cell.column #stores columns of dates in global dict. Columns same for all sheets
                cell.fill = PatternFill("solid", fgColor="FFEFC2")
                if datetime.strftime(datetime.strptime(cell.value, '%d %b %Y'), '%a').startswith('S'): #weekend
                    cell.fill = PatternFill("solid", fgColor="AFD5FF")
            start_date = start_date + timedelta(days=1)
        # excel_schedule_workbook.save(str(workbook_location))

def clear_future_columns(start_date:datetime):
    get_date_row()
    for i in excel_schedule_workbook.sheetnames:
        ws = excel_schedule_workbook[i]
        ws.delete_cols(date_columns[datetime.strftime(start_date, '%d %b %Y')], ws.max_column)
    build_date_row()
    get_date_row()

def get_date_row():
    current_ws = excel_schedule_workbook['TSE1']
    for cell in current_ws[date_row]:
        date_columns[str(cell.value)] = cell.column #stores columns of dates in global dict. Columns same for all sheets

# returns dict with Worksheet as key and an array of names as value
def get_all_names():
    for sheet in excel_schedule_workbook.worksheets:
        get_all_wiw_users()
        sheet_names = {} # names : column number
        for cell in sheet['A']:
            sheet_names[cell.value] = cell.row
            if cell.value in all_wiw_full_names:
                all_wiw_full_names[cell.value] = 'active'
        all_names[sheet.title] = sheet_names

def populate_excel_sheet(shifts_in):
    for user_id in shifts_in:
        user = get_user_from_id(user_id)
        all_to_requests = get_time_off_requests_for_user(user_id)
        user_shifts = shifts_in[user_id]
        location_ids = [user_shifts[0].location_id]
        schedule_name = check_location_id(user_shifts[0].location_id)
        for shift in user_shifts:
            if shift.location_id not in location_ids:
                location_ids.append(shift.location_id)
        if schedule_name == 0:
            continue
        check_sheet_for_name(user.full_name, schedule_name)
        populate_user_in_excel_sheet(user_shifts, schedule_name, user)
        try:
            if len(all_to_requests[user_id]) > 0:
                populate_users_time_off(all_to_requests[user_id], location_ids, user)
        except Exception as e:
            # print(e.__cause__)
            continue

def check_sheet_for_name(name_in, schedule_name):
    if name_in not in all_names[schedule_name]:
        #add to worksheet
        ws = excel_schedule_workbook[schedule_name]
        current_row = ws.max_row+1
        if current_row == 2:
            current_row = 3
        ws.cell(row=current_row, column=1).value = name_in
        #add to dict
        all_names[schedule_name][name_in] = current_row
        
def populate_user_in_excel_sheet(user_shifts, schedule_name, user):
        team_number = False
        for shift in user_shifts:
           try:
                ws = excel_schedule_workbook[check_location_id(shift.location_id)]  
                check_sheet_for_name(user.full_name, check_location_id(shift.location_id))     
                current_cell = ws.cell(row=all_names[check_location_id(shift.location_id)][user.full_name], column=date_columns[datetime.strftime(shift.start_time, '%d %b %Y')])
                active_cell = ws.cell(row=current_cell.row, column=2)
                try:
                    active_cell.value = all_wiw_full_names[user.full_name]
                except Exception as e:
                    active_cell.value = 'inactive'
                    _ = e
           except KeyError as e:
                # print("line 306" + str(e))
                continue

           current_cell.value = shift.length
           if shift.location_id not in [5227330,5233779] and 0 <= int(datetime.strftime(shift.start_time, '%-H')) <= 6:
               current_cell.value = str(int(current_cell.value)) + 'N'
           elif shift.location_id not in [5227330,5233779]:
                current_cell.value = str(int(current_cell.value)) + 'D'
            
           start_to_end = ': ' + str(datetime.strftime(shift.start_time, '%-H')) + ' - ' + str(datetime.strftime(shift.end_time, '%-H'))
           current_cell.value = str(current_cell.value) + start_to_end

           if shift.published == True:
                current_cell.fill = PatternFill("solid", fgColor=shift.color)
           else:
                current_cell.fill = PatternFill("solid", fgColor='ffffff')
                current_cell.border = Border(left=Side(border_style='thick', color=shift.color),right=Side(border_style='thick', color=shift.color),top=Side(border_style='thick', color=shift.color),bottom=Side(border_style='thick', color=shift.color))
        #    current_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
           current_cell.comment = Comment(shift.notes, "iSOC Scheduling")

def populate_users_time_off(user_requests, location_ids, user):
    if user.full_name == '':
        _=0 #Used for debugging

    for location in location_ids:
        schedule_name = check_location_id(location)
        for request in user_requests:
            if request.user_status in [1,4]:
                continue
            ws = excel_schedule_workbook[schedule_name]
            date_check = request.start_time
            while request.start_time <= date_check < request.end_time:
                try:
                    current_cell = ws.cell(row=all_names[schedule_name][user.full_name], column=date_columns[datetime.strftime(date_check, '%d %b %Y')])
                except KeyError as e:
                    date_check = date_check + timedelta(days=1)
                    continue                
                current_cell.value = 'V - Time Off'
                current_cell.fill = PatternFill("solid", fgColor='ff8789')
                current_cell.comment = Comment(request.type_label, "iSOC Scheduling")
                date_check = date_check + timedelta(hours=24)

# ------------------------- WHEN I WORK METHODS --------------------------- #

def get_team_number(site_id):
    try:
        n = int(site_id)
    except:
        print("Team number not INT")
    teams = {
        4781862:1 ,4781869:2,4781872:3,4781873:4,4781874:5,4781875:6,4781876:7,4781877:8,4781878:9,4781879:10, 0:0
     }
    try:
         return teams[n]
    except: 
        return site_id

#takes in location (schedule) ID and returns which worksheet
def check_location_id(location_id) -> str:
    schedules = {
        #5129876 : "Default"
        5132409 : "TSE1",
        5132410 : "TSE2",
        5134192 : "TSE3",
        5132412 : "TechOps",
        # 5189759 : "Colin Test",
        5227330 : "EMEATier1",
        5331591 : "EMEATier1",
        5231727 : "ShiftLeads",
        5233779 : "EMEATier3"
    }
    try:
        return schedules[int(location_id)]
    except Exception as e:
        return 0

# gets all shifts as shift classes and stores in dictionary by user
def get_all_shifts():    
    url_headers = get_url_and_headers('shifts?start=' + str(datetime(2022, 3, 1)) + "&end=" + str(datetime.now()+ timedelta(days=360)) + "&unpublished=true")
    response = requests.request("GET", url_headers[0], headers=url_headers[1])
    all_shifts = response.json()['shifts']
    employee_shifts = {}
    for i in all_shifts:
        start_time = datetime.strptime(i['start_time'], '%a, %d %b %Y %H:%M:%S %z').astimezone(pytz.timezone('UTC'))
        end_time = datetime.strptime(i['end_time'], '%a, %d %b %Y %H:%M:%S %z').astimezone(pytz.timezone('UTC'))
        new_shift = classes.shift(int(i['id']), int(i['account_id']), int(i['user_id']), int(i['location_id']), int(i['position_id']),
                                int(i['site_id']), start_time, end_time, bool(i['published']), bool(i['acknowledged']), i['notes'], i['color'], bool(i['is_open']))
        if int(i['user_id']) in employee_shifts:
            current_users_shifts = employee_shifts.get(int(i['user_id']))
            current_users_shifts.append(new_shift)
            employee_shifts[int(i['user_id'])] = current_users_shifts
        else: 
            employee_shifts[int(i['user_id'])] = [new_shift]
    return employee_shifts

def get_time_off_requests_for_user(user_id):
    url_headers = get_url_and_headers('requests?start=' + str(datetime(2022, 3, 1)) + "&end=" + str(datetime.now()+timedelta(days=100)) + '&user_id=' + str(user_id) + '&limit=200')
    try: #permissions error
        response = requests.request("GET", url_headers[0], headers=url_headers[1])
        all_requests = response.json()['requests']
    except Exception as e:
        all_requests = {}
    return store_time_off(all_requests)

def store_time_off(all_requests):
    # creates a unique hash for each shift by multiplying the start time by user ID
    # allows for quick lookup of duplicate shifts
    def get_shift_hash(shift_in):
        start_time = int(shift_in.start_time.strftime('%Y%m%d%H%M'))
        user_id = int(shift_in.user_id)
        end_time = int(shift_in.end_time.strftime('%Y%m%d%H%M'))
        try:
            user_status = int(shift_in.user_status)
            return start_time * user_id * user_status
        except Exception as e:
            return start_time * user_id

    requests = {}
    hashed_requests = {}
    for i in all_requests:
        start_time = datetime.strptime(i['start_time'], '%a, %d %b %Y %H:%M:%S %z').astimezone(pytz.timezone('UTC'))
        end_time = datetime.strptime(i['end_time'], '%a, %d %b %Y %H:%M:%S %z').astimezone(pytz.timezone('UTC'))
        new_request = classes.time_off_request(i['id'], i['account_id'],i['user_id'], i['status'],i['type'],i['type_id'], i['hours'], start_time, end_time, 0,i['user_status'],i['type_label'])
        shift_hash = get_shift_hash(new_request)
        if new_request.user_status not in [0,2,3]:
            continue
        if shift_hash in hashed_requests:
            if new_request.to_id == hashed_requests[shift_hash].to_id:
                continue
            # delete_time_off(new_request, token)
            continue
        else:
            hashed_requests[shift_hash] = new_request
        if int(i['user_id']) in requests:
            current_users_requests =requests.get(int(i['user_id']))
            current_users_requests.append(new_request)
            requests[int(i['user_id'])] = current_users_requests
        else: 
            requests[int(i['user_id'])] = [new_request]
        if len(requests[int(i['user_id'])]) > 199:
            print("TOO MANY TO REQUESTS for " + i['user_id'])
            break
    # Store by Date!
    return requests

def get_all_wiw_users():
    all_users = {}
    url_headers = get_url_and_headers('users')
    response = requests.request("GET", url_headers[0], headers=url_headers[1])
    all_users_json = response.json()['users']
    for u in all_users_json:
        this_user = classes.user(u['first_name'], u['last_name'], u['email'], u['id'], u['positions'], u['role'], u['locations'], u['is_hidden'], u['is_active'])
        all_users[u['id']] = (this_user)
        try:
            if all_wiw_full_names[this_user.full_name] == 'active':
                continue
        except Exception as e:
            all_wiw_full_names[this_user.full_name]='inactive' 
            _=e
    return all_users

#takes in user_id and returns user object
def get_user_from_id(user_id):
    url_headers = get_url_and_headers('users/'+str(user_id))
    i = 1
    while i < 10:
        try:
            j = requests.request("GET", url_headers[0], headers=url_headers[1]).json()['user']
            break
        except:
            i +=1 
    return classes.user(j['first_name'], j['last_name'], j['email'], j['id'],j['positions'],j['role'],j['locations'],j['is_hidden'],j['is_active'])


def main():
    build_date_row()
    clear_future_columns(datetime.now()-timedelta(days=14))
    get_all_names()
    populate_excel_sheet(get_all_shifts())
    excel_schedule_workbook.save(str(authentication.get_schedule_location()))


if __name__ == "__main__":
    main()