#!/usr/bin/env python3
""""
Author: Colin McAllister (https://github.com/offsetkeyz)

This script will pull all shifts from When I Work and will put them into an excel spreadsheet.
"""

import json
import os
import sys
from openpyxl import load_workbook

token = ''

# Open the configuration json to access API login credentials
path_to_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config-colin.json')
if not os.path.isfile(path_to_json):
    sys.exit("'config.json' not found! Please add it and try again.")
else:
    with open(path_to_json) as file:
        config = json.load(file)

workbook_location = '/Users/colin.mcallister/Library/CloudStorage/OneDrive-ArcticWolfNetworksInc/Documents/triage_schedule_from_wiw.xlsx'
iSOC_Team_Structure = load_workbook('/Users/colin.mcallister/Library/CloudStorage/OneDrive-ArcticWolfNetworksInc/Documents/iSOC Team Structure - Colin.xlsx', data_only=True)
weekly_headcount = load_workbook('/Users/colin.mcallister/Downloads/Weekly Headcount Report - S2.xlsx', data_only=True)
workbook = load_workbook(workbook_location)
date_columns = {}
all_names = {}

home_office_override = {
    '2549928': 'ON', #Alex Zidik
    '2549904' : 'ON' #James Hryszko
}

all_positions = {
    'Business Analyst 1':10762839,
    'Business Systems Manager':10762840,
    'Co-op/ Intern':10762841,
    'Concierge Sec Eng 3':10762842,
    'Concierge Sec Engineer 2':10762843,
    'Concierge Sec Engnr 2': 10762843,
    'Concierge Security Eng 2':10762843,
    'Concierge Sec Engnr 3':10762842,
    'Concierge Security Eng 3':10762842,
    'Dir, Security Services':10762845,
    'IT Applications Admin':10762846,
    'Mgr Concierge Services':10762847,
    'Mgr Technical Operations':10762848,
    'Mgr, Concierge Security':10762847,
    'Mgr, Security Operations':10762849,
    'Network Ops Supp Analyst':10762850,
    'Ntwk and Tech Ops Analyst':10762850,
    'Network Sec Ops Analyst 2':10762850,
    'Network Sec Ops Anlst 1':10762850,
    'Network Sec Ops Analyst 1':10762850,
    'S2 Technical Trainer 4':10762851,
    'Shift Lead Sec Ops':10762852,
    'Shift Lead Security Ops':10762852,
    'Sr Dir, Sec Ops':10762854,
    'Sr Dir Security Ops':10762854,
    'Sr Mgr, Security Ops':10762855,
    'SVP, Security Services':10762856,
    'Tech Lead Security Svcs':10762857,
    'Triage Sec Eng 1':10762858,
    'Triage Security Engr 1':10762858,
    'Triage Security Eng 1':10762858,
    'Triage Security Analyst':10762858,
    'Triage Security Eng 2':10762859,
    'Triage Sec Eng 2':10762859,
    'Triage Security Engr 2':10762859,
    'Triage Sec Eng 3':10762860,
    'Triage Security Eng 3':10762860,
    'Triage Security Engr 3':10762860,
    'Triage Security Engr 4':10762861,
    'Triage Sec Eng 4':10762861,
    'VP, Business Applications':10762862,
    'CAN': 10762911,
    'USA': 10762910,
    'DEU':10762912,
    'DNK':10762912,
    'GBR':10762912,
    'DEU': 10824067,
    'FL': 10824070,
    'MN': 10824070,
    'ON': 10824070, #EST
    'TX': 10824071, #CST
    'UK': 10824072,
    'UT': 10824073 #MST
}

def list_all_wiw_users(token):
    all_users = []
    url_headers = bs_methods.get_url_and_headers('users', token)
    response = requests.request("GET", url_headers[0], headers=url_headers[1])
    for i in response.json()['users']:
        all_users.append(shift_classes.user(i['first_name'], i['last_name'], i['email'], i['employee_code'], i['positions'], i['role'], i['locations'], i['is_hidden'], i['is_active']))
    return all_users