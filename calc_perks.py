#!/usr/bin/env python3
""""
Author: Colin McAllister (https://github.com/offsetkeyz)

This will calculate perks based on the excel schedule.
"""

from datetime import timedelta
import datetime
import random

from time import strptime
from tracemalloc import start
from cogs import authentication, classes
from openpyxl import load_workbook
import datetime
from dateutil.tz import *

schedule_wb = load_workbook(authentication.get_schedule_location())
perks_workbook = load_workbook(authentication.get_perks_sheet_location())
perks_ws = perks_workbook['iSOC Perks']
start_date = ''
end_date = ''

all_perks_names = []
all_perks_cells = []
all_isoc_names = {}
all_new_employees = {}

CA_stat_nights= ['11 Nov 2022','12 Nov 2022', '24 Dec 2022', '25 Dec 2022', '26 Dec 2022','27 Dec 2022','31 Dec 2022','01 Jan 2023','02 Jan 2023','03 Jan 2023', '20 Feb 2023', '21 Feb 2023']
CA_stat_dates = ['11 Nov 2022', '24 Dec 2022', '25 Dec 2022', '26 Dec 2022', '31 Dec 2022','01 Jan 2023','02 Jan 2023', '20 Feb 2023']
US_stat_dates = ['24 Nov 2022', '25 Nov 2022', '23 Dec 2022','24 Dec 2022','25 Dec 2022','26 Dec 2022','31 Dec 2022','01 Jan 2023','02 Jan 2023','16 Jan 2023', '20 Feb 2023']
US_stat_nights = ['24 Nov 2022', '25 Nov 2022', '26 Nov 2022', '23 Dec 2022','24 Dec 2022','25 Dec 2022','26 Dec 2022','27 Dec 2022','31 Dec 2022','01 Jan 2023','02 Jan 2023','03 Jan 2023','16 Jan 2023','17 Jan 2023','20 Feb 2023','21 Feb 2023']


def get_start_date():
    input_date = input().strip()
    try:
        dt_start = datetime.datetime.strptime(str(input_date), '%d %b %Y')
        dt_start = dt_start.replace(hour=13)
    except ValueError as e:
        print ("Incorrect format with: " + input_date)
        print('Please try again (ex. 07 Jul 2033): ')
        input_date = get_start_date()
    return dt_start

#strips leading and trailing spaces
def strip_spaces(name_in):
    if type(name_in) is not str:
        return name_in
    if name_in != None:
        name_in_stripped = name_in.strip()
        return name_in_stripped 
    return name_in

def find_canadian_row():
    CA_limit_row = 125
    for cell in perks_ws['A']:
        if cell.value =='Subtotal CA Hourly':
            return cell.row
    print("error with find_canadian_row line 55")
    return CA_limit_row

def is_canadian(employee_in, CA_limit_row):
    if find_name_in_perks(employee_in.name) < CA_limit_row:
        return True
    return False

# Add numbers into correct column. Function takes name, night hours, 12H day hours
    #search for name and return row number. (cell.row)
def find_name_in_perks(name_in) -> int:
    for i in range(len(all_perks_names)):
        if strip_spaces(name_in) == strip_spaces(all_perks_names[i]):
            return all_perks_cells[i].row
    return -1

def compare_names():
    for i in range(0,5,1):
        for cell in schedule_wb.worksheets[i]['A']:
            try:
                if len(cell.value) > 1 and '>' not in str(cell.value):
                    all_isoc_names[cell.value] = schedule_wb.worksheets[i].cell(row=cell.row, column=2).value
            except Exception as e:
                continue
    for cell in perks_ws['D']:
        try: 
            active_cell = perks_ws.cell(row=cell.row, column=2)
            if len(cell.value) > 1 and '>' not in str(cell.value) and all_isoc_names[cell.value]=='active':
                all_perks_names.append(cell.value)
                all_perks_cells.append(cell)
        except Exception as e:
            continue
    absent_names = []
    names_to_remove = []
    for name in all_isoc_names:
        if all_isoc_names[name] == 'inactive' and name in all_perks_names:
            names_to_remove.append(name)
            continue
        if name not in all_perks_names and all_isoc_names[name] == 'active':
            absent_names.append(name)

    if len(names_to_remove) > 0:
        print(f'Please remove from perks template: {names_to_remove} --------------------')
    if len(absent_names)>0:
        print(f'{absent_names} need to be added to perks template.')
        print('re-run this script once everyone has been added')
        exit()

def calc_percs_by_section(CA_limit_row, save_location, start_date, end_date):
    all_names = {} # names : column number
    date_columns = {}

    #stores columns of dates in global dict. Columns same for all sheets
    current_ws = schedule_wb['TSE1']
    for cell in current_ws[2]:
        date_columns[str(cell.value)] = cell.column 

    for sheet in schedule_wb.worksheets:
        sheet_names = {}
        for cell in sheet['A']:
            if cell.coordinate not in ['A1','A2']:                
                sheet_names[cell.value] = cell.row
        all_names[sheet.title] = sheet_names
        for row in sheet.iter_rows(3, sheet.max_row, min_col=date_columns[datetime.datetime.strftime(start_date, '%d %b %Y')], max_col=date_columns[datetime.datetime.strftime(end_date, '%d %b %Y')]):
            new_employee_name = sheet.cell(row=row[0].row, column = 1).value
            if (new_employee_name in all_new_employees) and (new_employee_name != None):
                new_employee = all_new_employees[new_employee_name]
                # print(str(new_employee_name) + 'already in schedule with shifts (line 163)')
            else:
                new_employee = classes.employee(name=new_employee_name, night_hours=0, weekend_hours=0, meals=0,overtime=0, row=row[0].row, shifts_worked={}, stat_hours_in=0)
                all_new_employees[new_employee_name] = new_employee  
   
            for cell in row:
                cell_value = cell.value
                check_and_calc_for_12s(new_employee, cell_value)
                if cell_value == None:
                    cell_value =0
                new_employee.shifts_worked[sheet.cell(row=2, column=cell.column).value] = cell_value
            calculate_stat(new_employee, CA_limit_row)
            add_to_perks(new_employee)
        print(f'Calculating {sheet.title}...')
    perks_workbook.save(save_location)
    print(f'File saved as cSOC_Perks_{str(start_date.strftime("%b%-d"))}-{end_date.strftime("%b%-d")}.xlsx')

 #used by above function. this does the counting    
def check_and_calc_for_12s(new_employee, cell_value):
    if str(cell_value).strip().startswith("12D"):
        new_employee.weekend_hours = new_employee.weekend_hours + 12
        new_employee.meals = new_employee.meals + 1
    if str(cell_value).strip().startswith("12N"):
        new_employee.night_hours = new_employee.night_hours + 12
        new_employee.meals = new_employee.meals + 1
    elif str(cell_value).__contains__('N') and len(str(cell_value)) <=3:
        if len(str(cell_value)) == 3:
            new_employee.night_hours = new_employee.night_hours + int(cell_value[:2])
        else:
            new_employee.night_hours = new_employee.night_hours + int(cell_value[0])

def calculate_stat(employee_in, CA_limit_row):
    if is_canadian(employee_in, CA_limit_row):
        
        for i in CA_stat_dates:
            current_value = str(employee_in.shifts_worked.get(i))
            employee_stat = employee_in.stat_hours
            hours_worked = cell_hours_worked(current_value)
            employee_stat += hours_worked
            employee_in.stat_hours = employee_stat
            # stat has been calculated so it removes this value before calculating OT
            employee_in.shifts_worked[i] = 0


            for j in CA_stat_nights:
                current_value = str(employee_in.shifts_worked.get(j))
                if '12N' in current_value:
                    employee_stat = employee_in.stat_hours
                    employee_stat += cell_hours_worked(current_value)
                    employee_in.stat_hours = employee_stat
                    employee_in.shifts_worked[j] = 0
    else: #if US employee
        
        for i in US_stat_dates:

            current_value = str(employee_in.shifts_worked.get(i))
            employee_stat = employee_in.stat_hours
            hours_worked = cell_hours_worked(current_value)
            employee_stat += hours_worked
            employee_in.stat_hours = employee_stat
            employee_in.shifts_worked[i] = 0

            for j in US_stat_nights:
                current_value = str(employee_in.shifts_worked.get(j))

                if '12N' in current_value:
                    employee_stat = employee_in.stat_hours
                    employee_stat += cell_hours_worked(current_value)
                    employee_in.stat_hours = employee_stat
                    employee_in.shifts_worked[i] = 0

# given the contents of the cell, returns an INT with how many hours worked
def cell_hours_worked(cell_value):
    if cell_value == None:
        return 0
    cell_value = strip_spaces(cell_value)
    if cell_value == '':
        return 0
    cell_value = str(cell_value)
    if 'V' in str(cell_value):
        return 0
    if 'S' in str(cell_value):
        return 0
    if cell_value[0].isdigit():
        if len(cell_value) > 1:
            if cell_value[1].isdigit():
                try:
                    return int(cell_value[:2])
                except:
                    print('error with cell_hours_worked')
        try:
            return int(cell_value[0])
        except:
            print('error with cell_hours_worked')
    return 0

def add_to_perks(new_employee): 
    # add data to perks WS
    if new_employee.name is not None:
        perks_row = find_name_in_perks(new_employee.name)
        if perks_row < 0:
            return

        current_cell = perks_ws.cell(row=perks_row, column=5)
        current_cell.value = new_employee.night_hours
        current_cell = perks_ws.cell(row=perks_row, column=7)
        current_cell.value = new_employee.weekend_hours

        current_cell = perks_ws.cell(row=perks_row, column=9)
        current_cell.value = new_employee.meals

        current_cell = perks_ws.cell(row=perks_row, column=12)
        current_cell.value = new_employee.stat_hours


def main():
    CA_limit_row = find_canadian_row()
    # token = authentication.authenticate_WiW_API()
    compare_names()

    print('Enter Start Date UTC (ex. 04 Feb 2030): ')
    start_date = get_start_date()
    print('Enter End Date: ')
    end_date = get_start_date()
    save_location = f'{str(authentication.get_perks_directory())}cSOC_Perks_{str(start_date.strftime("%b%-d"))}-{end_date.strftime("%b%-d")}.xlsx'
    print(f'I am robot and I will calculate perks for the period {str(start_date.strftime("%b%-d"))}-{end_date.strftime("%b%-d")}... Is this correct? (Y/N): ')
    r = input()
    if r.strip().lower() not in ['y', 'yes']:
        print("ok bye")
        return
    print(random.choice(["lets get perking...","$$$", "I hope I'm doing this right..."]))
    calc_percs_by_section(CA_limit_row, save_location, start_date, end_date)

if __name__ == "__main__":
    main()
    
